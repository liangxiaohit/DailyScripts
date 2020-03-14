import shutil
import os
import openpyxl
from MainFiles.DateReader import Reader
#导入其他类，需要写整个路径MainFiles.DateReader.


class MainPy:

    FormPath = r"E:\\Work\\定时配送\\Form\\"
    FormName = "Form.xlsx"
    FormNameNW = "Form1.xlsx"

    NJKPath = r"E:\\Work\\定时配送\\NJK\\2020"
    NSPath = r"E:\\Work\\定时配送\\New Save\\2020"
    NWPath = r"E:\\Work\\定时配送\\New world\\2020"
    JYPath = r"E:\\Work\\定时配送\\YA\\2020"

    NJKShowWords = "NJK DELIVERY FORM"
    NSShowWords = "New Save DELIVERY FORM"
    JYShowWords = "CHILLI YA DELIVERY FORM"
    NWShowWords = "New World DELIVERY FORM"

    formula_NJK = "=B8*5+B9*8+B10*8+B11*8+B12*8+B13*10"
    formula_NS = "=B8*5+B9*8+B10*8+B11*8+B12*8+B13*10"
    formula_JY = "=B8*5+B9*8+B10*8+B11*8+B12*8+B13*10"
    formula_NW = "没有公式，不需要传"

    path_date = None
    week_days = None
    def __init__(self):
        print("MainPy is run")

    def getDateForWeek(self):
        Dr = Reader()
        path_date = Dr.creat_date()
        MainPy.path_date = str(path_date)
        MainPy.week_days = Dr.week_days

    def creatNewFiles(self):
        # pass
        MainPy.creatWb(self, MainPy.NJKPath, MainPy.FormPath, MainPy.FormName, MainPy.NJKShowWords)
        MainPy.creatWb(self, MainPy.NSPath, MainPy.FormPath, MainPy.FormName, MainPy.NSShowWords)
        MainPy.creatWb(self, MainPy.JYPath, MainPy.FormPath, MainPy.FormName, MainPy.JYShowWords)
        MainPy.creatWb(self, MainPy.NWPath, MainPy.FormPath, MainPy.FormNameNW, MainPy.NWShowWords)

    # 该函数传入3个变量：1，njk到2020的目录，2，初始表格的目录，3，初始表格的名字（NW与其他表格略有不同）4,exls的名字
    def creatWb(self, path_folder, form_path, form_name, file_name):

        # 查看是否存在该星期属于的月份，如果不存在则创建，存在则略过
        if MainPy.checkIfNeedCreatMonthFolder(self, MainPy.path_date, path_folder + "\\"):
            pass
        else:
            os.makedirs(path_folder + "\\" + MainPy.path_date[0:2] + "月")
        # 在月的文件夹下创建星期的文件夹
        os.makedirs(path_folder + "\\" + MainPy.path_date[0:2] + "月" + "\\" + MainPy.path_date)
        # 拷贝Form.xlsx到各个子目录中
        createdFilePath = path_folder + "\\" + MainPy.path_date[0:2] + "月" + "\\" + MainPy.path_date
        shutil.copy(form_path + form_name, createdFilePath)
        # 将文件重命名
        original_Name = createdFilePath + "\\" + form_name
        new_Name = createdFilePath + "\\" + MainPy.path_date + " " + file_name + ".xlsx"
        os.rename(original_Name, new_Name)

        MainPy.modifyFile(self, new_Name, file_name)

    def checkIfNeedCreatMonthFolder(self, path_date, check_path)-> bool:
        isFolderExist = False
        month_word = str(path_date)[0:2]
        dir_content = os.listdir(check_path)
        print(dir_content)
        for i in range(len(dir_content)):
            print(dir_content[i][0:2])
            if month_word == dir_content[i][0:2]:
                isFolderExist = True

        print(isFolderExist)
        return isFolderExist


    def modifyFile(self, file, title_name):

        excel = openpyxl.load_workbook(file)
        sheets = excel.sheetnames
        for i in range(sheets.__len__()):
            ws = excel[sheets[i]]
            if i != 7:
                if title_name.__eq__(MainPy.NWShowWords):
                    ws.title = str(MainPy.week_days[i])
                else:
                    ws.title = str(MainPy.week_days[i])
                    ws["H4"] = str(MainPy.week_days[i])
            else:
                ws.title = "总计"
                ws["B5"] = str(MainPy.week_days[0])
                ws["C5"] = str(MainPy.week_days[1])
                ws["D5"] = str(MainPy.week_days[2])
                ws["E5"] = str(MainPy.week_days[3])
                ws["F5"] = str(MainPy.week_days[4])
                ws["G5"] = str(MainPy.week_days[5])
                ws["H5"] = str(MainPy.week_days[6])

            ws.cell(1,1).value = title_name


        excel.save(file)
        excel.close()

main_Ins = MainPy()
main_Ins.getDateForWeek()
MainPy.creatNewFiles(main_Ins)